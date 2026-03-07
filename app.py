from io import BytesIO
import re

import openpyxl
import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from openpyxl.worksheet.formula import ArrayFormula

from ycharts.exceptions import YChartsAPIError, YChartsAuthError
from ycharts.fetcher import fetch_latest_value

load_dotenv()

DATA_START_ROW = 19

st.set_page_config(page_title="YCharts Template Populator", layout="centered")

st.title("YCharts Template Populator")
st.write("Upload a holdings file (like Zabbour.xlsx) and a Template workbook. Holdings populate from row 19.")

holdings_file = st.file_uploader("1) Upload Holdings File", type=["xlsx"])
template_file = st.file_uploader("2) Upload Template Workbook", type=["xlsx"])

YCHARTS_FUNC_PATTERN = re.compile(r"=\s*(?:_xll\.)?(?:@)?(?P<func>YC\w+)\((?P<args>.*)\)\s*$", re.IGNORECASE)
CELL_REF_PATTERN = re.compile(r"^\$?[A-Za-z]{1,3}\$?\d+$")


def _split_formula_args(arg_text: str) -> list[str]:
    args: list[str] = []
    current = []
    in_quotes = False
    for ch in arg_text:
        if ch == '"':
            in_quotes = not in_quotes
            current.append(ch)
        elif ch == "," and not in_quotes:
            args.append("".join(current).strip())
            current = []
        else:
            current.append(ch)
    if current:
        args.append("".join(current).strip())
    return args


def _resolve_argument(ws, raw_arg: str):
    arg = raw_arg.strip()
    if arg.startswith('"') and arg.endswith('"') and len(arg) >= 2:
        return arg[1:-1]
    if CELL_REF_PATTERN.match(arg):
        return ws[arg.replace("$", "")].value
    return arg


def _extract_formula_text(value) -> str | None:
    if isinstance(value, ArrayFormula):
        return value.text
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _compute_ycharts_ticker(symbol) -> str | None:
    if symbol is None:
        return None
    s = str(symbol).strip().upper()
    if not s:
        return None
    if s == "GOOGL":
        return "GOOGL"
    if s == "YAHOO":
        return "YAHOO"
    if s == "INDVMUNI":
        return "MUB"
    if len(s) > 4 and s.endswith("X"):
        return f"M:{s}"
    if len(s) > 8:
        return "BND"
    return s


def _fetch_ycharts_value(ticker, metric, func_name: str) -> float | None:
    if ticker is None or metric is None:
        return None

    ticker_text = str(ticker).strip().upper()
    metric_text = str(metric).strip().lower()
    if not ticker_text or not metric_text:
        return None

    if ticker_text.startswith("M:") or func_name.upper() in {"YCF", "YCSMF"}:
        return fetch_latest_value(ticker_text.replace("M:", ""), metric_text, "fund")
    return fetch_latest_value(ticker_text, metric_text, "security")


def _get_holdings_rows(workbook) -> list[list[object]]:
    ws = workbook.worksheets[0]
    rows: list[list[object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = list(row[:5])
        if any(v is not None and str(v).strip() != "" for v in values):
            rows.append(values)
    return rows


if holdings_file is not None and template_file is not None:
    with st.spinner("Loading workbooks..."):
        holdings_file.seek(0)
        holdings_wb = openpyxl.load_workbook(holdings_file, data_only=True)

        template_file.seek(0)
        template_wb = openpyxl.load_workbook(template_file, data_only=False)

    if "Template" not in template_wb.sheetnames:
        st.error("Template workbook is missing a sheet named `Template`.")
    else:
        template_ws = template_wb["Template"]
        holdings_rows = _get_holdings_rows(holdings_wb)

        if not holdings_rows:
            st.error("No holdings rows found in uploaded holdings file.")
        else:
            for i, row_values in enumerate(holdings_rows):
                target_row = DATA_START_ROW + i
                for col_idx, val in enumerate(row_values, start=1):
                    template_ws.cell(row=target_row, column=col_idx, value=val)
                template_ws.cell(row=target_row, column=6, value=_compute_ycharts_ticker(row_values[2]))

            start_row = DATA_START_ROW
            end_row = DATA_START_ROW + len(holdings_rows) - 1

            progress = st.progress(0)
            results: list[dict[str, object]] = []
            errors: list[dict[str, str]] = []

            ycharts_cells: list[tuple[str, str, object, object]] = []
            for row in template_ws.iter_rows(min_row=start_row, max_row=end_row, min_col=7, max_col=template_ws.max_column):
                for cell in row:
                    formula_text = _extract_formula_text(cell.value)
                    if not formula_text:
                        continue
                    match = YCHARTS_FUNC_PATTERN.match(formula_text.strip())
                    if not match:
                        continue
                    func_name = match.group("func").upper()
                    args = _split_formula_args(match.group("args"))
                    if len(args) < 2:
                        continue
                    ticker = _resolve_argument(template_ws, args[0])
                    metric = _resolve_argument(template_ws, args[1])
                    ycharts_cells.append((cell.coordinate, func_name, ticker, metric))

            if not ycharts_cells:
                st.warning(
                    f"No YCharts formulas found between rows {start_row} and {end_row}. "
                    "The template uses formulas like =_xll.YCI(...) / =_xll.YCP(...)."
                )
            else:
                total = len(ycharts_cells)
                for idx, (coord, func_name, ticker, metric) in enumerate(ycharts_cells, start=1):
                    try:
                        value = _fetch_ycharts_value(ticker, metric, func_name)
                        template_ws[coord] = value
                        results.append(
                            {
                                "Cell": f"Template!{coord}",
                                "Function": func_name,
                                "Ticker": ticker,
                                "Metric": metric,
                                "Value": value,
                            }
                        )
                    except YChartsAuthError as exc:
                        errors.append({"Cell": f"Template!{coord}", "Error": f"Auth error: {exc}"})
                    except YChartsAPIError as exc:
                        errors.append({"Cell": f"Template!{coord}", "Error": f"API error: {exc}"})
                    progress.progress(idx / total)

                out = BytesIO()
                template_wb.save(out)
                out.seek(0)

                st.success(
                    f"Inserted {len(holdings_rows)} holdings rows ({start_row}-{end_row}) and "
                    f"resolved {len(results)} YCharts formula cells."
                )
                st.download_button(
                    label="Download Filled Template",
                    data=out.getvalue(),
                    file_name="Template_Filled.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                if results:
                    st.subheader("Resolved Values")
                    st.dataframe(pd.DataFrame(results), use_container_width=True)

                if errors:
                    st.subheader("Errors")
                    st.dataframe(pd.DataFrame(errors), use_container_width=True)
